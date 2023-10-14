#Importing all the necessary libraries
import time
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select

#Initialising WebDriver
service_obj = Service("C:\Drivers\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj)
driver.implicitly_wait(10)

driver.get("https://www.moneycontrol.com/fixed-income/calculator/axis-bank/fixed-deposit-calculator-UTI10-BAB001.html#result")
driver.maximize_window()

#Selecting the XLSx file where the test data and results are available
file="F:\Selenium_projects\Interest Checking Project\Interest Calculator Testing.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Interest Calculator"]

#Getting the total number of rows and columns
rows=sheet.max_row


for row in range(2, rows+1):
    #Passing data from each row in the application
    principle=sheet.cell(row,1).value
    rate_of_interest=sheet.cell(row,2).value
    period=sheet.cell(row,3).value
    period_type=sheet.cell(row,4).value
    frequency=sheet.cell(row,5).value
    maturity_value=sheet.cell(row,6).value

    #Sending the vlues to input tag
    driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(principle)
    driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(rate_of_interest)
    driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(period)

    select_period_type=Select(driver.find_element(By.XPATH, "//select[@id='tenurePeriod']"))
    select_period_type.select_by_visible_text(period_type)
    select_frequency=Select(driver.find_element(By.XPATH, "//select[@id='frequency']"))
    select_frequency.select_by_visible_text(frequency)
    driver.find_element(By.XPATH, "//div[@class='cal_div']//a[1]").click()
    
    time.sleep(2)
    result=driver.find_element(By.CSS_SELECTOR, "span[id='resp_matval'] strong").text

    #Validation
    if float(result)==float(maturity_value):
        print("Test Pass!")
        sheet.cell(row,8).value="Passed!"
        green_fill=PatternFill(start_color="60b212",
                               end_color="60b212",
                               fill_type="solid")
        sheet.cell(row,8).fill=green_fill
    else:
        print("Test Failed!")
        sheet.cell(row,8).value="Failed!"
        red_fill=PatternFill(start_color="ff0000",
                               end_color="ff0000",
                               fill_type="solid")
        sheet.cell(row,8).fill=red_fill

    driver.find_element(By.XPATH, "//img[@class='PL5']").click()
    time.sleep(1)

workbook.save(file)
driver.close()